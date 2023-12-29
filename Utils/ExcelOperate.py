import openpyxl
from openpyxl import load_workbook

from ConfigFiles.ConfigPath import xlsxPath


class ExcelOperate:
    def __init__(self):
        self.workbook = None
        self.sheet = None

    def load_workbook(self,filename):
        '''
        加载相应的excel文件
        :param filename:
        :return:
        '''
        try:
            self.workbook = load_workbook(filename)
        except Exception as e:
            print(e)

    def get_sheet_by_name(self,sheetname):
        '''
        拿到xlsx文件里对应的页
        :param sheetname:
        :return:
        '''
        try:
            self.sheet = self.workbook[sheetname]
        except Exception as e:
            print(e)

    def  get_rows_nums(self):
        '''
        返回当前页的最大行数
        :return:
        '''
        return self.sheet.max_row
    def get_col_nums(self):
        '''
        获取最大列数
        :return:
        '''
        return self.sheet.max_column

    def get_row_values(self,row):
        '''
        根据row获取某一行的值
        :param row:
        :return:
        '''
        columns = self.sheet.max_column
        row_data = []
        #遍历列的时候从1开始，不是从0开始，因为xlsx没有第0行，第0列
        for i in range(1,columns+1):
            cell_value = self.sheet.cell(row = row,column = i).value
            row_data.append(cell_value)
        return row_data

    def get_cell_value(self,row,column):
        '''
        获取某一个单元格的值
        :param row:
        :param column:
        :return:
        '''
        cell_value = self.sheet.cell(row= row,column=column).value
        return cell_value

if __name__ == '__main__':
    eo = ExcelOperate()
    #根据名字读取xlsx文件，xlsxPath已经在ConfigPath.py文件中定义过了
    eo.load_workbook(xlsxPath)
    eo.get_sheet_by_name('Sheet1')
    print(eo.get_col_nums())
    print(eo.get_rows_nums())
    print(eo.get_row_values(1))
    print(eo.get_cell_value(2,3))



